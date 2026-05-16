# Instance Var --> Declare in constructor ie. __init__ method --> Call using self. keyword
# Class Var --> Decalre in Parent class --> Call using name of class.
from selenium.webdriver.common.by import By
import openpyxl


class HomePage():
    shop = (By.XPATH, "//a[contains(@href, 'shop')]")

    #Automated Form's objects
    name = (By.NAME, "name")
    email = (By.NAME, "email")
    password = (By.ID, "exampleInputPassword1")
    checkbox = (By.ID, "exampleCheck1")
    gender = (By.ID, "exampleFormControlSelect1")
    submit = (By.XPATH, "//input[@type='submit']")
    radio_button = (By.ID, "inlineRadio2")
    success_banner = (By.CLASS_NAME, "alert.alert-success.alert-dismissible")

    def __init__(self, driver):
        self.driver = driver

    def shopItems(self):
        return self.driver.find_element(*HomePage.shop)            # * -- to run it sequentially in tuple

    def getName(self):
        return self.driver.find_element(*HomePage.name)

    def getEmail(self):
        return self.driver.find_element(*HomePage.email)

    def getPassword(self):
        return self.driver.find_element(*HomePage.password)

    def clickCheckbox(self):
        return self.driver.find_element(*HomePage.checkbox).click()

    def selectGender(self):
        return self.driver.find_element(*HomePage.gender)

    def submitForm(self):
        return self.driver.find_element(*HomePage.submit).click()

    def employmentStatus(self):
        return self.driver.find_element(*HomePage.radio_button).click()

    def successBannerMsg(self):
        return self.driver.find_element(*HomePage.success_banner).text


# In python, to call the method which we declare in class, firstly we create object of the class then we call method like object.method name
# but if you want to call the method using classname.method then mark that method as @staticmethod so that you will not required to create the object of class
# For static method self parameter is not required
    @staticmethod
    def getExcelData(test_case_name):
        excel_file = openpyxl.load_workbook("/Users/pkolhe/Documents/Python_DataDriven.xlsx")
        sheet = excel_file.active
        Dict = {}
        for r in range(1, sheet.max_row + 1):
            if sheet.cell(row=r, column=1).value == test_case_name:
                for c in range(2, sheet.max_column + 1):  # Starting from column 2 because I dont want to print first name column
                    # load the excel data in dictionary --> Dict["Key"]="Value"
                    Dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value  # Row=1 because header row will be constant
                return [Dict]       # printes list of dict because we pass parameter in fixture method as a list