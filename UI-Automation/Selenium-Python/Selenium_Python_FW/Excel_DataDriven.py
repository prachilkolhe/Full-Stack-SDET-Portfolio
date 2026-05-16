import openpyxl

# Local path of excel file
excel_file = openpyxl.load_workbook("/Users/pkolhe/Documents/Python_DataDriven.xlsx")

# Selected the active sheet from the excel file
sheet = excel_file.active
Dict= {}        #Initiated a dictionary
# Reading the values from excel file
cell_value = sheet.cell(row=1, column=2)
print(cell_value.value)
print(sheet['B1'].value)            # can use cell number directly

# To write back in excel file
sheet.cell(row=2, column=2).value = "Virat"
print(sheet.cell(row=2, column=2).value)

# To print maximum rows and columns present in file
print(sheet.max_row)
print(sheet.max_column)

# Use the debugger to know how loops are working
# To print all the values of col.1 in each row
for i in range(1, sheet.max_row+1):             # +1 -- because it will excludes the last row so avoid we added it by 1
    print(sheet.cell(row=i, column=1).value)

# To print all the values of all columns and rows
for r in range(1, sheet.max_row+1):
    for c in range(1, sheet.max_column+1):
        print(sheet.cell(row=r, column=c).value)

# Use debugger to know the workflow of loop, also while debugging to know the what value is storing in variable select it while debugging is running, click right click and select option Evaluate expression
# To print the only of test case 2
for a in range(1, sheet.max_row+1):
    if sheet.cell(row=a, column=1).value == "Testcase 2":
        for b in range(2, sheet.max_column+1):          # Starting from column 2 because I dont want to print first name column
            # print(sheet.cell(row=a, column=b).value)
            # To load the excel data in dictionary --> Dict["Key"]="Value"
            Dict[sheet.cell(row=1, column=b).value] = sheet.cell(row=a, column=b).value      # Row=1 because header row will be constant
print(Dict)
